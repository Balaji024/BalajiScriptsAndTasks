import openpyxl
import os
os.chdir('H:\pythonScripts\BalajiScriptsAndTasks\PythonScripts')


#to load a Excel file to perform any operations on it
workbook = openpyxl.load_workbook('workbook.xlsx')
print(type(workbook), "1")

# to get particular sheet in workbook
sheet= workbook.get_sheet_by_name('Sheet1')
print(type(sheet), "2")

# to get sheets in workbook
sheets= workbook.get_sheet_names()
print(sheets,"3")

# to get cells in sheet
print(sheet['A1'].value, "5")

print (sheet.cell(row=1, column=2), "6")

for i in range(1,3):
    print(i, sheet.cell(row=i, column=2).value,"7")


